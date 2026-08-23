# SEO記事別 月次実績をExcelに出力する
# 指標1〜3はダッシュボードの data/dashboard.json から、指標4はGA4から新規取得
import sys, json, datetime as dt
from pathlib import Path
sys.path.insert(0, str(Path.home() / "dev/jagoo-marketing-dashboard/scripts"))
import collect_all as ca
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(sys.argv[1])
D = json.load(open(Path.home() / "dev/jagoo-marketing-dashboard/data/dashboard.json"))
months = D["months"]
start, end = months[0] + "-01", (dt.date.today() - dt.timedelta(days=1)).isoformat()
arts, af, ag, kws = D["articles"], D["article_funnel"], D["article_gsc"], D["article_keywords"]

# --- 指標4: 記事着地セッションのうち 資料DLページ到達 / DL完了ページ到達（GA4新規取得）
ga4 = ca.GA4(ca.cred_json("GA4_CREDENTIALS_JSON", "ga4_jagoo_oauth"))
def by_landing(extra):
    out = {}
    for d, mt in ga4.rows(start, end, ["yearMonth", "landingPage"], ["sessions"],
                          ca.f_and(ca.COLUMN_ARTICLES, extra)):
        p, ym = ca.to_path(d[1]), ca.ym_key(d[0])
        out.setdefault(p, {})[ym] = out.get(p, {}).get(ym, 0) + int(mt[0])
    return out
land_doc = by_landing(ca.f_contains("pagePath", "/document/"))
land_cmp = by_landing(ca.f_contains("pagePath", "complete/"))
print("GA4取得: DL到達 %d記事 / DL完了 %d記事" % (len(land_doc), len(land_cmp)))

gsc_m = {}  # 記事→月→[clicks, impr]
for p, days in ag.items():
    for day, (c, i) in days.items():
        ym = day[:7]
        v = gsc_m.setdefault(p, {}).setdefault(ym, [0, 0]); v[0] += c; v[1] += i

paths = sorted(set(arts) | set(ag) | set(af), key=lambda p: -sum(arts.get(p, {}).values()))

def g(dic, p, m, idx=None):
    v = dic.get(p, {}).get(m)
    if v is None: return 0
    return v[idx] if idx is not None else v

SHEETS = [
 ("1_セッション数", "セッション数（GA4・記事着地セッション・国=日本）", "count",
  lambda p, m: g(arts, p, m)),
 ("2_GSC表示回数", "GSC表示回数", "count", lambda p, m: g(gsc_m, p, m, 1)),
 ("2_GSCクリック数", "GSCクリック数", "count", lambda p, m: g(gsc_m, p, m, 0)),
 ("3_記事→DL遷移率", "記事→資料DLページ遷移率（％）＝記事から/document/へ遷移したPV÷記事PV（ダッシュボードと同じ計算）", "rate",
  lambda p, m: (g(af, p, m).get("to_doc", 0) if g(af, p, m) else 0, g(af, p, m).get("pv", 0) if g(af, p, m) else 0)),
 ("4_DL遷移→完了転換率", "DL遷移→DL完了転換率（％）＝記事着地セッションのうちDL完了ページ到達数÷資料DLページ到達数（GA4新規集計）", "rate",
  lambda p, m: (g(land_cmp, p, m), g(land_doc, p, m))),
 ("4a_DLページ到達S数", "記事着地セッションのうち資料DLページ(/document/)到達セッション数（転換率の分母）", "count", lambda p, m: g(land_doc, p, m)),
 ("4b_DL完了S数", "記事着地セッションのうちDL完了ページ(complete/)到達セッション数（転換率の分子）", "count", lambda p, m: g(land_cmp, p, m)),
]

wb = Workbook(); wb.remove(wb.active)
hdr_fill = PatternFill("solid", fgColor="DDDDDD"); bold = Font(bold=True)
for name, title, kind, fn in SHEETS:
    ws = wb.create_sheet(name)
    ws.append([title]); ws["A1"].font = bold
    ws.append(["期間: %s 〜 %s（%s は月途中）" % (months[0], months[-1], months[-1])])
    ws.append([])
    head = ["メインの注力キーワード", "記事URL"] + months + ["合計"]
    ws.append(head)
    for c in ws[4]: c.font = bold; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
    tot = {m: [0, 0] for m in months}
    for p in paths:
        row = [kws.get(p, ""), "https://jagoo.co.jp" + p + "/"]
        if kind == "count":
            vals = [fn(p, m) for m in months]
            for m, v in zip(months, vals): tot[m][0] += v
            row += vals + [sum(vals)]
        else:
            n_all = d_all = 0
            for m in months:
                n, d = fn(p, m); n_all += n; d_all += d; tot[m][0] += n; tot[m][1] += d
                row.append(round(n / d * 100, 2) if d else None)
            row.append(round(n_all / d_all * 100, 2) if d_all else None)
        ws.append(row)
    # 合計行
    if kind == "count":
        tr = ["全記事 合計", ""] + [tot[m][0] for m in months]; tr.append(sum(tr[2:]))
    else:
        tr = ["全記事 合計", ""] + [round(tot[m][0] / tot[m][1] * 100, 2) if tot[m][1] else None for m in months]
        N, Dn = sum(t[0] for t in tot.values()), sum(t[1] for t in tot.values())
        tr.append(round(N / Dn * 100, 2) if Dn else None)
    ws.append(tr)
    for c in ws[ws.max_row]: c.font = bold
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 52
    for i in range(3, len(head) + 1): ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = "C5"
    if kind == "rate":
        for r in ws.iter_rows(min_row=5, min_col=3):
            for c in r: c.number_format = "0.00"

ws = wb.create_sheet("読み方", 0)
for line in [
 "SEO記事別 月次実績（ダッシュボード「SEO流入詳細」タブ準拠）",
 "出力日: %s / データ期間: %s 〜 %s" % (dt.date.today().isoformat(), start, end),
 "",
 "・記事の対象: URLに /column/ を含むページ。GA4は国=日本で絞り込み（ダッシュボードと同じ）",
 "・1 セッション数: その記事に着地（最初に見た）したセッション数",
 "・2 GSC表示回数/クリック数: Search Consoleのページ別日次データを月で合計",
 "・3 記事→DL遷移率: ダッシュボードと同じ計算（記事から/document/へ遷移したPV ÷ 記事PV）。移動元が取れない閲覧があるため低めに出る近似値",
 "・4 DL遷移→完了転換率: 今回新規に集計。その記事に着地したセッションのうち、資料DLページ(/document/)に到達した数を分母、DL完了ページ(complete/)に到達した数を分子にした割合",
 "  ※3と4は計算単位が異なる（3はPV・移動元ベース、4はセッション・着地記事ベース）ため、掛け算で記事→完了率にはならない",
 "  ※4a・4bシートに分母・分子の実数を入れている。件数が少ない記事は転換率がぶれるので実数と合わせて見ること",
 "・合計列: 件数は13ヶ月の合計、率は分子合計÷分母合計",
 "・注力キーワード: シート「SEO記事アクセス計測」から抽出した対応表（未登録の記事は空欄）",
]: ws.append([line])
ws.column_dimensions["A"].width = 120
wb.save(OUT); print("保存:", OUT, "記事数:", len(paths))
